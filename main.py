import os
import sys
import json
import queue
import threading
import webbrowser
import tkinter as tk
from tkinter import ttk, filedialog, messagebox


import argparse
import html
import json
import sys
from pathlib import Path

# ---------------------------------------------------------------------------
# 1. BÍ DANH KHÓA — giúp thích nghi khi tên trường thay đổi giữa các file
# ---------------------------------------------------------------------------
ALIASES = {
    "tieu_de":       ["tieu_de", "title", "tieude", "ten", "heading"],
    "ngay_cap_nhat": ["ngay_cap_nhat", "ngay", "date", "updated", "ngaycapnhat"],
    "don_vi_ban":    ["don_vi_ban", "nha_cung_cap", "seller", "shop", "cong_ty", "donvi"],
    "bang_gia":      ["bang_gia", "banggia", "prices", "price_list", "products", "san_pham_list"],
    "ghi_chu_gia":   ["ghi_chu_gia", "ghi_chu", "note", "notes", "ghichu"],
    "lien_he":       ["lien_he", "lienhe", "contact", "contacts"],
}
PROD_ALIASES = {
    "san_pham": ["san_pham", "sanpham", "ten", "name", "product", "loai"],
    "gia":      ["gia", "price", "don_gia", "dongia", "value"],
    "don_vi":   ["don_vi", "donvi", "unit", "dvt", "dvt_tinh"],
}
CONTACT_ALIASES = {
    "so_dien_thoai": ["so_dien_thoai", "sdt", "phone", "phones", "dien_thoai", "hotline"],
    "dia_chi":       ["dia_chi", "diachi", "address", "addresses", "dc"],
    "email":         ["email", "emails", "mail"],
    "website":       ["website", "web", "url", "trang_web"],
}


def pick(d, aliases, default=None):
    """Lấy giá trị đầu tiên khớp một trong các bí danh (không phân biệt hoa/thường)."""
    if not isinstance(d, dict):
        return default
    lower = {k.lower(): v for k, v in d.items()}
    for a in aliases:
        if a.lower() in lower and lower[a.lower()] not in (None, ""):
            return lower[a.lower()]
    return default


def as_list(v):
    """Ép về list: None->[], scalar->[scalar], list giữ nguyên."""
    if v is None:
        return []
    if isinstance(v, list):
        return [x for x in v if x not in (None, "")]
    return [v]


# ---------------------------------------------------------------------------
# 2. GIẢI NÉN — bóc mọi lớp chuỗi-chứa-JSON cho tới khi ra object/list thật
# ---------------------------------------------------------------------------
def deep_parse(node):
    """Trả về list các dict bản ghi, dù node lồng sâu tới đâu."""
    out = []

    def walk(x):
        if x is None:
            return
        if isinstance(x, str):
            s = x.strip()
            if not s:
                return
            try:
                walk(json.loads(s))     # chuỗi có thể chứa JSON -> bóc tiếp
            except (json.JSONDecodeError, ValueError):
                return                  # chuỗi thường, không phải bản ghi -> bỏ
        elif isinstance(x, list):
            for item in x:
                walk(item)
        elif isinstance(x, dict):
            out.append(x)

    walk(node)
    return out


def normalize(rec):
    """Đưa một dict thô về cấu trúc chuẩn để render."""
    prices = []
    for p in as_list(pick(rec, ALIASES["bang_gia"], [])):
        if not isinstance(p, dict):
            continue
        prices.append({
            "san_pham": str(pick(p, PROD_ALIASES["san_pham"], "") or ""),
            "gia":      str(pick(p, PROD_ALIASES["gia"], "") or ""),
            "don_vi":   str(pick(p, PROD_ALIASES["don_vi"], "") or ""),
        })

    lh_raw = pick(rec, ALIASES["lien_he"], {}) or {}
    lien_he = {
        "so_dien_thoai": [str(x) for x in as_list(pick(lh_raw, CONTACT_ALIASES["so_dien_thoai"], []))],
        "dia_chi":       [str(x) for x in as_list(pick(lh_raw, CONTACT_ALIASES["dia_chi"], []))],
        "email":         [str(x) for x in as_list(pick(lh_raw, CONTACT_ALIASES["email"], []))],
        "website":       str(pick(lh_raw, CONTACT_ALIASES["website"], "") or ""),
    }

    return {
        "tieu_de":       str(pick(rec, ALIASES["tieu_de"], "(không tiêu đề)") or "(không tiêu đề)"),
        "ngay_cap_nhat": str(pick(rec, ALIASES["ngay_cap_nhat"], "") or ""),
        "don_vi_ban":    str(pick(rec, ALIASES["don_vi_ban"], "") or ""),
        "bang_gia":      prices,
        "ghi_chu_gia":   str(pick(rec, ALIASES["ghi_chu_gia"], "") or ""),
        "lien_he":       lien_he,
    }

"""
Tìm kiếm Google qua SerpApi và gộp kết quả nhiều trang.
Khóa API được TRUYỀN VÀO, không hardcode trong mã nguồn.
"""
import time

try:
    import serpapi
except ImportError:
    serpapi = None


def fetch_all_pages(keyword, api_key, num_pages=10,
                    location="Viet Tri, Phu Tho Province, Vietnam",
                    progress=None, stop_flag=None):
    """
    Gọi SerpApi start=0,10,... rồi gộp organic_results (đã loại trùng link).

    progress(msg): hàm callback nhận chuỗi tiến trình (tùy chọn).
    stop_flag(): hàm trả True nếu người dùng yêu cầu dừng (tùy chọn).
    """
    if serpapi is None:
        raise RuntimeError(
            "Chưa cài thư viện serpapi. Chạy: pip install google-search-results")
    if not api_key:
        raise ValueError("Thiếu SerpApi key.")

    def log(msg):
        if progress:
            progress(msg)

    client = serpapi.Client(api_key=api_key)
    all_results, seen = [], set()

    for page in range(num_pages):
        if stop_flag and stop_flag():
            log("Đã dừng theo yêu cầu.")
            break

        params = {
            "engine": "google",
            "q": keyword,
            "location": location,
            "google_domain": "google.com.vn",
            "hl": "vi",
            "gl": "vn",
            "start": page * 10,
            "num": 10,
            "api_key": api_key,
        }
        print (params)
        try:
            data = client.search(params)
        except Exception as e:                       # noqa: BLE001
            log(f"  Trang {page+1}: lỗi gọi API ({e}) — dừng.")
            break

        if "error" in data:
            log(f"  Trang {page+1}: dừng ({data['error']})")
            break

        organic = data.get("organic_results", [])
        if not organic:
            log(f"  Trang {page+1}: hết kết quả — dừng.")
            break

        new = [r for r in organic if r.get("link") not in seen]
        for r in new:
            seen.add(r.get("link"))
        all_results.extend(new)
        log(f"  Trang {page+1}: +{len(new)} (tổng {len(all_results)})")

        time.sleep(1)  # lịch sự với API

    return all_results
"""
Module dùng chung: nhận danh sách organic_results (đã gộp từ nhiều trang),
trích xuất thông tin và xuất ra 1 file Excel gồm:
  - Sheet "Ket qua": mỗi dòng 1 kết quả (STT, Trang/Site, Tiêu đề, Nơi bán, Link, SĐT, Vị trí, Đánh giá)
  - Sheet "Thong ke": số kết quả theo từng site + biểu đồ cột
"""
import re
from urllib.parse import urlparse
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# Ký tự điều khiển bị cấm trong .xlsx (XML). Loại trước khi ghi ô.
_ILLEGAL_XLSX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')

def _safe(value):
    """Làm sạch giá trị trước khi ghi vào ô Excel."""
    if isinstance(value, str):
        return _ILLEGAL_XLSX.sub('', value)
    return value

# ------- Trích xuất số điện thoại Việt Nam trong text -------
PHONE_RE = re.compile(
    r'(?<!\d)(?:\+?84|0)(?:\d[\s.\-]?){8,10}\d(?!\d)'
)

def extract_phone(*texts):
    for t in texts:
        if not t:
            continue
        for m in PHONE_RE.finditer(t):
            raw = m.group(0)
            digits = re.sub(r'\D', '', raw)
            if digits.startswith('84'):
                digits = '0' + digits[2:]
            # SĐT VN hợp lệ: 10 số bắt đầu bằng 0, hoặc 11 số (cố định cũ)
            if digits.startswith('0') and len(digits) in (10, 11):
                return digits
    return ''

def get_domain(link):
    try:
        net = urlparse(link).netloc
        return net[4:] if net.startswith('www.') else net
    except Exception:
        return ''

def parse_results(organic_results):
    """Chuyển organic_results (list dict) -> list dòng dữ liệu đã trích xuất."""
    rows = []
    for r in organic_results:
        link = r.get('link', '')
        snippet = r.get('snippet', '')
        title = r.get('title', '')
        rich = r.get('rich_snippet', {}).get('top', {}).get('detected_extensions', {})
        rating = rich.get('rating', '')
        reviews = rich.get('reviews', '')
        danhgia = f"{rating} ({reviews})" if rating else ''
        rows.append({
            'stt': r.get('position', len(rows) + 1),
            'site': get_domain(link),
            'title': title,
            'noiban': r.get('source', ''),      # "source" của Google thường là tên nơi bán/thương hiệu
            'link': link,
            'sdt': extract_phone(snippet, title),
            'vitri': r.get('position', ''),
            'danhgia': danhgia,
        })
    return rows

# ------- Xuất Excel -------
HEADER_FILL = PatternFill('solid', start_color='1F4E78')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
CELL_FONT = Font(name='Arial', size=10)
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def build_excel(rows, keyword, out_path):
    wb = Workbook()

    # ---- Sheet 1: Ket qua (bang chinh duy nhat) ----
    ws = wb.active
    ws.title = "Ket qua"
    has_crawl = any('noidung' in r for r in rows)
    headers = ['STT', 'Trang (Site)', 'Tiêu đề', 'Nơi bán', 'Link',
               'Số điện thoại', 'Vị trí', 'Dữ liệu text rút gọn']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER

    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r['stt'])
        ws.cell(row=i, column=2, value=_safe(r['site']))
        ws.cell(row=i, column=3, value=_safe(r['title']))
        ws.cell(row=i, column=4, value=_safe(r['noiban']))
        c5 = ws.cell(row=i, column=5, value=_safe(r['link']))
        if r['link']:
            c5.hyperlink = r['link']
            c5.font = Font(name='Arial', size=10, color='0563C1', underline='single')
        # ưu tiên SĐT tìm được trên trang, nếu không có thì lấy từ snippet
        sdt = r.get('sdt_trang') or r.get('sdt', '')
        ws.cell(row=i, column=6, value=_safe(sdt))
        ws.cell(row=i, column=7, value=r['vitri'])
        ws.cell(row=i, column=8, value=_safe(r.get('text_rutgon', '')))
        for c in range(1, 9):
            cell = ws.cell(row=i, column=c)
            cell.border = BORDER
            if c != 5:
                cell.font = CELL_FONT
            cell.alignment = Alignment(vertical='top', wrap_text=(c in (3, 8)))

    widths = [6, 22, 40, 22, 46, 18, 8, 70]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = w
    ws.freeze_panes = 'A2'

    # ---- Sheet 2: Thong ke + bieu do ----
    ws2 = wb.create_sheet("Thong ke")
    counts = Counter(r['site'] for r in rows if r['site'])
    ws2.append(['Trang (Site)', 'Số kết quả'])
    for c in range(1, 3):
        cell = ws2.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = BORDER
    for site, n in counts.most_common():
        ws2.append([site, n])
    last = ws2.max_row
    total_row = last + 1
    ws2.cell(row=total_row, column=1, value='Tổng').font = Font(name='Arial', bold=True)
    ws2.cell(row=total_row, column=2, value=f'=SUM(B2:B{last})').font = Font(name='Arial', bold=True)
    for row in ws2.iter_rows(min_row=2, max_row=last, min_col=1, max_col=2):
        for cell in row:
            cell.font = CELL_FONT
            cell.border = BORDER
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 14

    if last >= 2:
        chart = BarChart()
        chart.type = "col"
        chart.title = f"Số kết quả theo trang - từ khóa: {keyword}"
        chart.y_axis.title = "Số kết quả"
        chart.x_axis.title = "Trang (Site)"
        data = Reference(ws2, min_col=2, min_row=1, max_row=last)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 9
        chart.width = 20
        chart.legend = None
        ws2.add_chart(chart, "D2")

    # ---- Sheet 3: thong tin tra cuu ----
    ws3 = wb.create_sheet("Thong tin")
    info = [
        ['Từ khóa tra cứu', keyword],
        ['Tổng số kết quả trích xuất', len(rows)],
        ['Số kết quả có SĐT', sum(1 for r in rows if r['sdt'])],
        ['Số trang (site) khác nhau', len(counts)],
    ]
    for row in info:
        ws3.append(row)
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 40
    for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(name='Arial', bold=True)

    # ---- Sheet 4: Noi dung (text da crawl) ----
    if has_crawl:
        ws4 = wb.create_sheet("Noi dung")
        h4 = ['STT', 'Trang (Site)', 'Link', 'Nội dung (text)']
        ws4.append(h4)
        for c in range(1, len(h4) + 1):
            cell = ws4.cell(row=1, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = BORDER
        for i, r in enumerate(rows, start=2):
            ws4.cell(row=i, column=1, value=r['stt'])
            ws4.cell(row=i, column=2, value=_safe(r['site']))
            ws4.cell(row=i, column=3, value=_safe(r['link']))
            ws4.cell(row=i, column=4, value=_safe(r.get('noidung', '')))
            for c in range(1, 5):
                cell = ws4.cell(row=i, column=c)
                cell.border = BORDER
                cell.font = CELL_FONT
                # cột nội dung: canh trên, KHÔNG wrap để không kéo giãn hàng quá cao
                cell.alignment = Alignment(vertical='top', wrap_text=False)
        ws4.column_dimensions['A'].width = 6
        ws4.column_dimensions['B'].width = 24
        ws4.column_dimensions['C'].width = 45
        ws4.column_dimensions['D'].width = 100
        ws4.freeze_panes = 'A2'

    wb.save(out_path)
    return out_path

"""
Crawl nội dung text từ các link kết quả.
Chỉ lấy TEXT (bỏ script, style, menu...), có timeout và User-Agent.
"""
import re
import time
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
import requests
from bs4 import BeautifulSoup

# Số luồng crawl song song. 8 là mức an toàn; tăng sẽ nhanh hơn
# nhưng dễ bị site chặn hoặc nghẽn mạng.
MAX_WORKERS = 8

# Khóa để in tiến trình không bị chèn lẫn giữa các luồng
_print_lock = threading.Lock()

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/122.0 Safari/537.36"
    ),
    "Accept-Language": "vi,en;q=0.8",
}

# Excel giới hạn ~32.767 ký tự / ô. Chừa biên an toàn.
MAX_CELL = 32000

# Regex SĐT Việt Nam (dùng lại để dò trong toàn văn trang)
PHONE_RE = re.compile(r'(?<!\d)(?:\+?84|0)(?:\d[\s.\-]?){8,10}\d(?!\d)')


# Ký tự điều khiển bị cấm trong file .xlsx (XML). Phải loại trước khi ghi Excel.
ILLEGAL_XLSX = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')


def strip_illegal(text):
    """Bỏ ký tự điều khiển mà Excel/XML không chấp nhận."""
    if not text:
        return text
    return ILLEGAL_XLSX.sub('', text)


def summarize_text(text, max_chars=1200):
    """
    Rút gọn text crawl: bỏ dòng rác/quá ngắn, ưu tiên câu có ý nghĩa,
    cắt <= max_chars ký tự để hiển thị gọn trong 1 ô Excel.
    """
    if not text:
        return ""
    lines = []
    for ln in text.splitlines():
        ln = ln.strip()
        # bỏ dòng quá ngắn (menu, nút bấm) và dòng chỉ có ký hiệu
        if len(ln) < 20:
            continue
        if not any(c.isalpha() for c in ln):
            continue
        lines.append(ln)
    joined = " ".join(lines)
    joined = re.sub(r'\s+', ' ', joined).strip()
    joined = strip_illegal(joined)
    if len(joined) > max_chars:
        joined = joined[:max_chars].rsplit(' ', 1)[0] + " ..."
    return joined


def clean_phones(text):
    """Trả về danh sách SĐT hợp lệ, không trùng, theo thứ tự xuất hiện."""
    found = []
    for m in PHONE_RE.finditer(text):
        digits = re.sub(r'\D', '', m.group(0))
        if digits.startswith('84'):
            digits = '0' + digits[2:]
        if digits.startswith('0') and len(digits) in (10, 11) and digits not in found:
            found.append(digits)
    return found


def extract_text(html):
    """Bỏ các thẻ không phải nội dung, lấy text sạch."""
    soup = BeautifulSoup(html, "html.parser")
    for tag in soup(["script", "style", "noscript", "svg", "nav", "footer",
                     "header", "form", "iframe"]):
        tag.decompose()
    # ưu tiên vùng nội dung chính nếu có
    main = soup.find("article") or soup.find("main") or soup.body or soup
    text = main.get_text(separator="\n")
    # gộp dòng trống, khoảng trắng thừa
    lines = [ln.strip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln]
    return strip_illegal("\n".join(lines))


def crawl_one(url, timeout=15):
    """
    Trả về dict: {ok, text, phones, error}
    text đã cắt <= MAX_CELL ký tự.
    """
    try:
        resp = requests.get(url, headers=HEADERS, timeout=timeout)
        resp.raise_for_status()
        # đảm bảo giải mã đúng tiếng Việt
        if not resp.encoding or resp.encoding.lower() == "iso-8859-1":
            resp.encoding = resp.apparent_encoding
        text = extract_text(resp.text)
        phones = clean_phones(text)
        truncated = len(text) > MAX_CELL
        if truncated:
            text = text[:MAX_CELL] + "\n...[đã cắt bớt]"
        return {"ok": True, "text": text, "phones": phones,
                "truncated": truncated, "error": ""}
    except requests.exceptions.Timeout:
        return {"ok": False, "text": "", "phones": [], "truncated": False,
                "error": "Timeout"}
    except requests.exceptions.RequestException as e:
        return {"ok": False, "text": "", "phones": [], "truncated": False,
                "error": str(e)[:200]}


def _process_one(idx, row, total):
    """Crawl 1 row, ghi kết quả vào chính row đó. Chạy trong 1 luồng."""
    url = row.get("link", "")
    if not url:
        row["noidung"] = ""
        row["text_rutgon"] = ""
        row["sdt_trang"] = ""
        row["crawl_status"] = "Không có link"
        with _print_lock:
            print(f"  [{idx+1}/{total}] (không có link) - bỏ qua")
        return

    res = crawl_one(url)
    row["noidung"] = res["text"]
    row["text_rutgon"] = summarize_text(res["text"])
    row["sdt_trang"] = ", ".join(res["phones"])
    row["crawl_status"] = "OK" if res["ok"] else res["error"]
    # nếu cột SĐT từ snippet trống mà trang có, điền vào
    if not row.get("sdt") and res["phones"]:
        row["sdt"] = res["phones"][0]
    with _print_lock:
        print(f"  [{idx+1}/{total}] {url[:60]} -> {row['crawl_status']}")


def crawl_rows(rows, max_workers=MAX_WORKERS):
    """
    Crawl SONG SONG bằng threading. Gắn vào mỗi row:
      row['noidung']  : text nội dung
      row['sdt_trang']: SĐT tìm thấy trên trang (nối bằng ', ')
      row['crawl_status']: 'OK' / lỗi

    Thứ tự trong list `rows` được GIỮ NGUYÊN vì mỗi luồng chỉ
    sửa đúng phần tử của mình (không append theo thứ tự hoàn thành).
    """
    total = len(rows)
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=max_workers) as pool:
        futures = [pool.submit(_process_one, i, r, total)
                   for i, r in enumerate(rows)]
        for f in as_completed(futures):
            # gọi result() để lỗi bên trong luồng (nếu có) được nêu ra
            f.result()
    ok = sum(1 for r in rows if r.get("crawl_status") == "OK")
    print(f"  Hoàn tất crawl: {ok}/{total} OK trong {time.time()-t0:.1f}s "
          f"({max_workers} luồng)")
    return rows


def load_records(paths):
    """Đọc 1..n file, giải nén, chuẩn hóa. Trả về (records, stats)."""
    records, raw_total = [], 0
    for path in paths:
        try:
            data = json.loads(Path(path).read_text(encoding="utf-8"))
        except Exception as e:                       # noqa: BLE001
            print(f"  ! Bỏ qua {path}: {e}", file=sys.stderr)
            continue
        found = deep_parse(data)
        raw_total += len(found)
        records.extend(normalize(r) for r in found)
    # bỏ bản ghi rỗng hoàn toàn (không tiêu đề, không giá, không người bán)
    records = [r for r in records
               if r["don_vi_ban"] or r["bang_gia"] or r["tieu_de"] != "(không tiêu đề)"]
    return records, raw_total


# ---------------------------------------------------------------------------
# 3. SINH HTML — cùng giao diện với bản trước, nhúng dữ liệu vào <script>
# ---------------------------------------------------------------------------
TEMPLATE = r"""<!DOCTYPE html>
<html lang="vi">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>__TITLE__</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Bricolage+Grotesque:opsz,wght@12..96,600;12..96,800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;700&display=swap');
  :root{--ink:#141210;--paper:#f6f2e9;--card:#fffdf7;--line:#ded5c2;--steel:#3b4a5a;--steel2:#5a7089;--rust:#b5502a;--rust-soft:#e9dcc9;--muted:#8a8071}
  *{box-sizing:border-box;margin:0;padding:0}
  body{background:var(--paper);color:var(--ink);font-family:'Inter',sans-serif;line-height:1.5;-webkit-font-smoothing:antialiased}
  .wrap{max-width:1180px;margin:0 auto;padding:0 24px}
  header{padding:56px 0 32px;border-bottom:2px solid var(--ink)}
  .eyebrow{font-family:'JetBrains Mono',monospace;font-size:12px;letter-spacing:.22em;text-transform:uppercase;color:var(--rust);font-weight:700;margin-bottom:14px}
  h1{font-family:'Bricolage Grotesque',sans-serif;font-weight:800;font-size:clamp(38px,6vw,72px);line-height:.98;letter-spacing:-.02em;max-width:16ch}
  .sub{margin-top:18px;color:var(--muted);max-width:56ch;font-size:16px}
  .stats{display:flex;margin-top:34px;border:1px solid var(--line);border-radius:10px;overflow:hidden;background:var(--card);flex-wrap:wrap}
  .stat{flex:1;min-width:120px;padding:18px 22px;border-right:1px solid var(--line)}
  .stat:last-child{border-right:none}
  .stat .n{font-family:'JetBrains Mono',monospace;font-weight:700;font-size:30px;color:var(--steel);line-height:1}
  .stat .l{font-size:12px;color:var(--muted);margin-top:6px}
  .toolbar{position:sticky;top:0;z-index:20;background:var(--paper);padding:20px 0 14px;margin-top:8px;border-bottom:1px solid var(--line);display:flex;gap:12px;flex-wrap:wrap;align-items:center}
  .search{flex:1;min-width:220px;position:relative}
  .search input{width:100%;padding:12px 14px 12px 40px;border:1px solid var(--line);border-radius:8px;background:var(--card);font-size:15px;font-family:inherit;color:var(--ink)}
  .search input:focus{outline:2px solid var(--steel2);outline-offset:-1px}
  .search svg{position:absolute;left:13px;top:50%;transform:translateY(-50%);color:var(--muted)}
  .count{font-family:'JetBrains Mono',monospace;font-size:13px;color:var(--muted);white-space:nowrap}
  .toggle{padding:11px 16px;border:1px solid var(--ink);border-radius:8px;background:var(--ink);color:var(--paper);font-family:inherit;font-size:13px;font-weight:600;cursor:pointer}
  .toggle:hover{background:var(--rust);border-color:var(--rust)}
  main{padding:28px 0 80px}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;margin-bottom:16px;overflow:hidden;transition:border-color .15s}
  .card:hover{border-color:var(--steel2)}
  .card-head{padding:20px 24px;cursor:pointer;display:flex;gap:18px;align-items:flex-start}
  .idx{font-family:'JetBrains Mono',monospace;font-weight:700;font-size:13px;color:var(--rust);padding-top:3px;min-width:32px}
  .ch-main{flex:1}
  .ch-title{font-family:'Bricolage Grotesque',sans-serif;font-weight:600;font-size:19px;letter-spacing:-.01em;line-height:1.25}
  .ch-meta{display:flex;gap:14px;flex-wrap:wrap;margin-top:9px;font-size:13px;color:var(--muted);align-items:center}
  .seller{color:var(--steel);font-weight:600}
  .pill{font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:700;padding:3px 9px;border-radius:20px}
  .pill.has{background:var(--rust-soft);color:var(--rust)}
  .pill.none{background:#eee8dc;color:var(--muted)}
  .date{font-family:'JetBrains Mono',monospace;font-size:12px}
  .chev{color:var(--muted);transition:transform .2s;padding-top:4px}
  .card.open .chev{transform:rotate(90deg)}
  .body{max-height:0;overflow:hidden;transition:max-height .3s ease}
  .card.open .body{max-height:6000px}
  .body-inner{padding:4px 24px 24px;border-top:1px solid var(--line)}
  .card.open .body-inner{padding-top:20px}
  table{width:100%;border-collapse:collapse;margin-top:4px;font-size:14px}
  th{text-align:left;font-family:'JetBrains Mono',monospace;font-size:11px;letter-spacing:.06em;text-transform:uppercase;color:var(--muted);padding:8px 12px;border-bottom:1px solid var(--line);font-weight:700}
  td{padding:10px 12px;border-bottom:1px solid #eee6d6}
  tr:last-child td{border-bottom:none}
  .price{font-family:'JetBrains Mono',monospace;font-weight:700;color:var(--steel);white-space:nowrap}
  .unit{color:var(--muted);font-size:12px}
  .prod{font-weight:500}
  .noprice{color:var(--muted);font-style:italic;font-size:14px;padding:6px 0}
  .note{margin-top:16px;padding:12px 16px;background:#f3ecdd;border-left:3px solid var(--rust);border-radius:0 6px 6px 0;font-size:13px;color:#5a5040}
  .contact{margin-top:16px;display:grid;grid-template-columns:repeat(auto-fit,minmax(200px,1fr));gap:10px 24px;font-size:13px}
  .contact .c-row{display:flex;gap:8px}
  .contact .c-k{font-family:'JetBrains Mono',monospace;font-size:11px;color:var(--muted);text-transform:uppercase;letter-spacing:.05em;min-width:52px;padding-top:2px}
  .contact a{color:var(--steel);text-decoration:none;border-bottom:1px solid var(--line)}
  .contact a:hover{color:var(--rust)}
  .empty{text-align:center;padding:60px 0;color:var(--muted)}
  footer{border-top:1px solid var(--line);padding:24px 0;font-size:12px;color:var(--muted);font-family:'JetBrains Mono',monospace}
  mark{background:#ffe9a8;color:inherit;border-radius:2px}
</style>
</head>
<body>
<header>
  <div class="wrap">
    <div class="eyebrow">Tổng hợp báo giá &middot; Vật liệu xây dựng</div>
    <h1>__H1__</h1>
    <p class="sub">Dữ liệu gộp từ file JSON, hiển thị rõ ràng theo từng nguồn: sản phẩm, đơn giá, ghi chú và liên hệ.</p>
    <div class="stats" id="stats"></div>
  </div>
</header>
<div class="wrap">
  <div class="toolbar">
    <div class="search">
      <svg width="16" height="16" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="7"/><path d="m21 21-4-4"/></svg>
      <input id="q" type="text" placeholder="Tìm theo nhà cung cấp, sản phẩm, tiêu đề...">
    </div>
    <span class="count" id="count"></span>
    <button class="toggle" id="toggleAll">Mở tất cả</button>
  </div>
</div>
<main><div class="wrap"><div id="list"></div></div></main>
<footer><div class="wrap">Nguồn: __SOURCE__ &middot; __N__ bản ghi hợp lệ</div></footer>
<script>
const DATA = __DATA__;
const list=document.getElementById('list'),q=document.getElementById('q'),countEl=document.getElementById('count');
let allOpen=false;
const totalProd=DATA.reduce((s,d)=>s+(d.bang_gia?d.bang_gia.length:0),0);
const withPrice=DATA.filter(d=>d.bang_gia&&d.bang_gia.length).length;
const sellers=new Set(DATA.map(d=>d.don_vi_ban).filter(Boolean)).size;
document.getElementById('stats').innerHTML=[[DATA.length,'Nguồn dữ liệu'],[totalProd,'Dòng báo giá'],[withPrice,'Nguồn có bảng giá'],[sellers,'Nhà cung cấp']].map(([n,l])=>`<div class="stat"><div class="n">${n}</div><div class="l">${l}</div></div>`).join('');
function esc(s){return (s||'').replace(/[&<>]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;'}[c]));}
function hl(s,t){s=esc(s);if(!t)return s;try{return s.replace(new RegExp('('+t.replace(/[.*+?^${}()|[\]\\]/g,'\\$&')+')','gi'),'<mark>$1</mark>');}catch(e){return s;}}
function render(term){
  term=(term||'').trim();const t=term.toLowerCase();let shown=0;list.innerHTML='';
  DATA.forEach((d,i)=>{
    const hay=(d.tieu_de+' '+d.don_vi_ban+' '+(d.bang_gia||[]).map(p=>p.san_pham).join(' ')).toLowerCase();
    if(t&&!hay.includes(t))return;shown++;
    const n=d.bang_gia?d.bang_gia.length:0;
    const card=document.createElement('div');card.className='card'+(allOpen?' open':'');
    let rows;
    if(n){rows=`<table><thead><tr><th>Sản phẩm</th><th>Giá</th><th>Đơn vị</th></tr></thead><tbody>`+d.bang_gia.map(p=>`<tr><td class="prod">${hl(p.san_pham,term)}</td><td class="price">${esc(p.gia)||'—'}</td><td class="unit">${esc(p.don_vi)}</td></tr>`).join('')+`</tbody></table>`;}
    else{rows=`<div class="noprice">Bản ghi này không có bảng giá (thường là bài kỹ thuật/trọng lượng).</div>`;}
    const lh=d.lien_he||{},parts=[];
    if(lh.so_dien_thoai&&lh.so_dien_thoai.length)parts.push(`<div class="c-row"><span class="c-k">ĐT</span><span class="c-v">${lh.so_dien_thoai.map(esc).join(' &middot; ')}</span></div>`);
    if(lh.dia_chi&&lh.dia_chi.length)parts.push(`<div class="c-row"><span class="c-k">Địa chỉ</span><span class="c-v">${lh.dia_chi.map(esc).join('<br>')}</span></div>`);
    if(lh.email&&lh.email.length)parts.push(`<div class="c-row"><span class="c-k">Email</span><span class="c-v">${lh.email.map(esc).join(', ')}</span></div>`);
    if(lh.website)parts.push(`<div class="c-row"><span class="c-k">Web</span><span class="c-v"><a href="${lh.website.startsWith('http')?lh.website:'https://'+lh.website}" target="_blank" rel="noopener">${esc(lh.website)}</a></span></div>`);
    const contact=parts.length?`<div class="contact">${parts.join('')}</div>`:'';
    const note=d.ghi_chu_gia?`<div class="note">${esc(d.ghi_chu_gia)}</div>`:'';
    const date=d.ngay_cap_nhat?`<span class="date">⟳ ${esc(d.ngay_cap_nhat)}</span>`:'';
    card.innerHTML=`<div class="card-head"><div class="idx">${String(i+1).padStart(2,'0')}</div><div class="ch-main"><div class="ch-title">${hl(d.tieu_de,term)}</div><div class="ch-meta"><span class="seller">${hl(d.don_vi_ban,term)||'—'}</span><span class="pill ${n?'has':'none'}">${n?n+' sản phẩm':'không giá'}</span>${date}</div></div><div class="chev"><svg width="18" height="18" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="m9 18 6-6-6-6"/></svg></div></div><div class="body"><div class="body-inner">${rows}${note}${contact}</div></div>`;
    card.querySelector('.card-head').addEventListener('click',()=>card.classList.toggle('open'));
    list.appendChild(card);
  });
  if(!shown)list.innerHTML=`<div class="empty">Không tìm thấy kết quả cho "<b>${esc(term)}</b>"</div>`;
  countEl.textContent=shown+' / '+DATA.length+' nguồn';
}
q.addEventListener('input',e=>render(e.target.value));
document.getElementById('toggleAll').addEventListener('click',function(){allOpen=!allOpen;this.textContent=allOpen?'Đóng tất cả':'Mở tất cả';render(q.value);});
render('');
</script>
</body>
</html>"""

"""
Trích xuất dữ liệu báo giá có cấu trúc từ nội dung crawl, dùng Gemini.
Dọn lại từ tien_su_ly.py: sửa lỗi đọc khóa dính '\\n', vòng lặp key,
và chuẩn hóa đầu ra thành list dict thay vì list chuỗi thô.
"""
import json
import logging
import concurrent.futures

try:
    from google import genai   # pip install google-generativeai
except ImportError:
    genai = None

logging.basicConfig(level=logging.INFO)

TIMEOUT_SECONDS = 120
MODELS = [
    "gemini-2.5-flash",
    "gemini-2.0-flash-lite",
    "gemini-2.5-flash-lite",
    "gemini-1.5-flash",
    "gemini-2.5-pro",
]

PROMPT = (
    "Bạn là công cụ trích xuất dữ liệu. Từ nội dung bài viết được cung cấp, "
    "hãy trích xuất thông tin và trả về DUY NHẤT một object JSON hợp lệ "
    "(không kèm giải thích, không markdown, không backtick).\n\n"
    "Cấu trúc JSON đầu ra:\n"
    "{\n"
    '  "tieu_de": "tiêu đề bài viết",\n'
    '  "ngay_cap_nhat": "ngày cập nhật giá nếu có",\n'
    '  "don_vi_ban": "tên công ty/nhà bán",\n'
    '  "bang_gia": [\n'
    '    {"san_pham": "tên sản phẩm", "gia": "giá", "don_vi": "đơn vị (vd: đồng/cây)"}\n'
    "  ],\n"
    '  "ghi_chu_gia": "ghi chú về giá nếu có",\n'
    '  "lien_he": {\n'
    '    "so_dien_thoai": [], "dia_chi": [], "email": [], "website": ""\n'
    "  }\n"
    "}\n\n"
    "Quy tắc: chỉ dùng thông tin có trong dữ liệu, không bịa. "
    'Trường thiếu để "" hoặc []. Giữ nguyên định dạng số của giá. '
    "Gộp số điện thoại, loại trùng. Trả JSON thuần bắt đầu bằng { kết thúc bằng }.\n\n"
    "Dữ liệu đầu vào:\n"
)


def load_api_keys(path):
    """Đọc danh sách khóa từ file, mỗi dòng một khóa, đã strip sạch."""
    with open(path, encoding="utf-8") as f:
        return [ln.strip() for ln in f if ln.strip()]


def _call(client, model, data_content):
    full = f"{PROMPT}\n\nDữ liệu JSON:\n{json.dumps(data_content, ensure_ascii=False)}"
    resp = client.models.generate_content(model=model, contents=full)
    return resp.text


def _clean_json(text):
    """Bóc backtick nếu model lỡ trả markdown, rồi parse."""
    if not text:
        return None
    t = text.strip()
    if t.startswith("```"):
        t = t.strip("`")
        if t[:4].lower() == "json":
            t = t[4:]
    t = t.strip()
    start, end = t.find("{"), t.rfind("}")
    if start == -1 or end == -1:
        return None
    try:
        return json.loads(t[start:end + 1])
    except json.JSONDecodeError:
        return None


def extract_one(api_keys, data_content, progress=None):
    """
    Thử lần lượt (khóa, model) cho tới khi ra JSON hợp lệ.
    Trả về dict đã parse, hoặc None nếu thất bại toàn bộ.
    """
    if genai is None:
        raise RuntimeError(
            "Chưa cài google-generativeai. Chạy: pip install google-generativeai")

    def log(m):
        if progress:
            progress(m)

    for api_key in api_keys:
        try:
            client = genai.Client(api_key=api_key)
        except Exception as e:                       # noqa: BLE001
            log(f"  Khóa lỗi, bỏ qua: {e}")
            continue

        for model in MODELS:
            try:
                with concurrent.futures.ThreadPoolExecutor() as ex:
                    fut = ex.submit(_call, client, model, data_content)
                    text = fut.result(timeout=TIMEOUT_SECONDS)
                parsed = _clean_json(text)
                if parsed:
                    log(f"  ✓ {model}")
                    return parsed
                log(f"  {model}: phản hồi không hợp lệ, thử tiếp.")
            except concurrent.futures.TimeoutError:
                log(f"  {model}: quá thời gian, bỏ qua.")
            except Exception as e:                   # noqa: BLE001
                log(f"  {model}: {e}")
    return None


def extract_all(api_keys, rows, progress=None, stop_flag=None):
    """
    Chạy trích xuất cho các row có 'noidung'. Trả về list dict kết quả.
    """
    jobs = [r for r in rows if r.get("noidung")]
    out = []
    total = len(jobs)
    for i, data_content in enumerate(jobs, 1):
        if stop_flag and stop_flag():
            if progress:
                progress("Đã dừng trích xuất theo yêu cầu.")
            break
        if progress:
            progress(f"[{i}/{total}] Đang trích xuất...")
        res = extract_one(api_keys, data_content, progress=progress)
        if res:
            out.append(res)
    return out


def build_html(records, title, source):
    data_js = json.dumps(records, ensure_ascii=False)
    return (TEMPLATE
            .replace("__TITLE__", html.escape(title))
            .replace("__H1__", html.escape(title))
            .replace("__SOURCE__", html.escape(source))
            .replace("__N__", str(len(records)))
            .replace("__DATA__", data_js))



def _lazy(name):
    """Import module lõi khi cần; báo lỗi rõ nếu thiếu thư viện ngoài."""
    import importlib
    try:
        return importlib.import_module(name)
    except ImportError as e:
        raise RuntimeError(
            f"Không import được '{name}': {e}.\n"
            "Có thể bạn chưa cài thư viện. Chạy:\n"
            "    pip install -r requirements.txt"
        ) from e


APP_TITLE = "Tra cứu & Báo giá"
DEFAULT_LOCATION = "Viet Tri, Phu Tho Province, Vietnam"


def safe_name(keyword):
    return "".join(c if c.isalnum() else "_" for c in keyword).strip("_") or "ket_qua"


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_TITLE)
        self.geometry("920x680")
        self.minsize(820, 600)

        # trạng thái chia sẻ giữa các luồng
        self.rows = []              # dữ liệu sau crawl
        self.extracted = []         # dữ liệu sau trích xuất Gemini
        self.log_q = queue.Queue()
        self._stop = threading.Event()
        self._busy = False

        self._build_style()
        self._build_ui()
        self.after(100, self._drain_log)

    # ---------- Giao diện ----------
    def _build_style(self):
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        style.configure("TButton", padding=6)
        style.configure("Accent.TButton", padding=8, font=("Segoe UI", 10, "bold"))
        style.configure("Header.TLabel", font=("Segoe UI", 14, "bold"))
        style.configure("Sub.TLabel", foreground="#666")

    def _build_ui(self):
        # Thanh tiêu đề
        top = ttk.Frame(self, padding=(16, 12))
        top.pack(fill="x")
        ttk.Label(top, text="Tool Tra Cứu + Báo Giá", style="Header.TLabel").pack(anchor="w")
        ttk.Label(top, text="Tìm → Crawl → Trích xuất → Xuất Excel/HTML",
                  style="Sub.TLabel").pack(anchor="w")

        nb = ttk.Notebook(self)
        nb.pack(fill="both", expand=True, padx=16, pady=(4, 8))
        self.tab_run = ttk.Frame(nb, padding=14)
        self.tab_cfg = ttk.Frame(nb, padding=14)
        self.tab_out = ttk.Frame(nb, padding=14)
        nb.add(self.tab_run, text="  Chạy quy trình  ")
        nb.add(self.tab_cfg, text="  Cấu hình & Khóa  ")
        nb.add(self.tab_out, text="  Xuất kết quả  ")

        self._build_config_tab()
        self._build_run_tab()
        self._build_output_tab()

        # Thanh trạng thái + tiến trình dưới cùng
        bottom = ttk.Frame(self, padding=(16, 6))
        bottom.pack(fill="x")
        self.progress = ttk.Progressbar(bottom, mode="indeterminate")
        self.progress.pack(side="left", fill="x", expand=True)
        self.status = ttk.Label(bottom, text="Sẵn sàng", style="Sub.TLabel")
        self.status.pack(side="right", padx=(10, 0))

    def _build_config_tab(self):
        f = self.tab_cfg
        # SerpApi keys file (giống cách nhập Gemini)
        ttk.Label(f, text="File khóa SerpApi (mỗi dòng 1 khóa)",
                  font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 2))
        self.var_serpfile = tk.StringVar(value=os.environ.get("SERPAPI_KEYS_FILE", ""))
        e1 = ttk.Entry(f, textvariable=self.var_serpfile, width=54)
        e1.grid(row=1, column=0, sticky="we")
        ttk.Button(f, text="Chọn file...", command=self._pick_serpfile).grid(
            row=1, column=1, sticky="e", padx=(8, 0))
        ttk.Label(f, text="Mỗi dòng 1 khóa SerpApi. Không lưu vào mã nguồn.",
                  style="Sub.TLabel").grid(row=2, column=0, sticky="w", pady=(2, 12))

        # Gemini keys file
        ttk.Label(f, text="File khóa Gemini (mỗi dòng 1 khóa)",
                  font=("Segoe UI", 10, "bold")).grid(row=3, column=0, sticky="w", pady=(0, 2))
        self.var_gemfile = tk.StringVar()
        e2 = ttk.Entry(f, textvariable=self.var_gemfile, width=54)
        e2.grid(row=4, column=0, sticky="we")
        ttk.Button(f, text="Chọn file...", command=self._pick_gemfile).grid(
            row=4, column=1, sticky="e", padx=(8, 0))
        ttk.Label(f, text="Tùy chọn — chỉ cần khi dùng bước Trích xuất Gemini.",
                  style="Sub.TLabel").grid(row=5, column=0, sticky="w", pady=(2, 12))

        # Thông số tìm kiếm
        ttk.Label(f, text="Số trang Google", font=("Segoe UI", 10, "bold")).grid(
            row=6, column=0, sticky="w")
        self.var_pages = tk.IntVar(value=10)
        ttk.Spinbox(f, from_=1, to=20, textvariable=self.var_pages, width=6).grid(
            row=6, column=1, sticky="w", pady=4)

        ttk.Label(f, text="Vị trí (location)", font=("Segoe UI", 10, "bold")).grid(
            row=7, column=0, sticky="w")
        self.var_loc = tk.StringVar(value=DEFAULT_LOCATION)
        ttk.Entry(f, textvariable=self.var_loc, width=40).grid(
            row=7, column=1, sticky="we", pady=4)

        ttk.Label(f, text="Số luồng crawl", font=("Segoe UI", 10, "bold")).grid(
            row=8, column=0, sticky="w")
        self.var_workers = tk.IntVar(value=8)
        ttk.Spinbox(f, from_=1, to=32, textvariable=self.var_workers, width=6).grid(
            row=8, column=1, sticky="w", pady=4)

        self.var_docrawl = tk.BooleanVar(value=True)
        ttk.Checkbutton(f, text="Crawl nội dung từng trang (tắt = chỉ lấy từ Google)",
                        variable=self.var_docrawl).grid(
            row=9, column=0, columnspan=2, sticky="w", pady=(8, 0))

        f.columnconfigure(0, weight=1)

    def _build_run_tab(self):
        f = self.tab_run
        row = ttk.Frame(f)
        row.pack(fill="x")
        ttk.Label(row, text="Từ khóa:", font=("Segoe UI", 10, "bold")).pack(side="left")
        self.var_kw = tk.StringVar(value="Sắt phi 10")
        ttk.Entry(row, textvariable=self.var_kw, width=40).pack(
            side="left", padx=(8, 8), fill="x", expand=True)

        btns = ttk.Frame(f)
        btns.pack(fill="x", pady=(10, 6))
        self.btn_run = ttk.Button(btns, text="▶  Chạy tìm + crawl",
                                  style="Accent.TButton", command=self.on_run_pipeline)
        self.btn_run.pack(side="left")
        self.btn_extract = ttk.Button(btns, text="✨  Trích xuất Gemini",
                                      command=self.on_extract, state="disabled")
        self.btn_extract.pack(side="left", padx=6)
        self.btn_stop = ttk.Button(btns, text="■  Dừng",
                                   command=self.on_stop, state="disabled")
        self.btn_stop.pack(side="left")
        self.btn_loadjson = ttk.Button(btns, text="Nạp JSON backup...",
                                       command=self.on_load_backup)
        self.btn_loadjson.pack(side="right")

        ttk.Label(f, text="Nhật ký:", style="Sub.TLabel").pack(anchor="w", pady=(6, 2))
        self.txt_log = tk.Text(f, height=18, wrap="word", bg="#1e1e1e", fg="#d4d4d4",
                               insertbackground="#d4d4d4", font=("Consolas", 9),
                               relief="flat")
        self.txt_log.pack(fill="both", expand=True)
        sb = ttk.Scrollbar(self.txt_log, command=self.txt_log.yview)
        self.txt_log.configure(yscrollcommand=sb.set)

    def _build_output_tab(self):
        f = self.tab_out
        ttk.Label(f, text="Sau khi có dữ liệu, xuất ra file:",
                  style="Sub.TLabel").pack(anchor="w", pady=(0, 10))

        g = ttk.Frame(f)
        g.pack(anchor="w")
        ttk.Button(g, text="📊  Xuất Excel (từ dữ liệu crawl)",
                   command=self.on_export_excel, width=36).grid(row=0, column=0, pady=5)
        ttk.Button(g, text="🌐  Xuất HTML tương tác (từ dữ liệu Gemini)",
                   command=self.on_export_html, width=42).grid(row=1, column=0, pady=5)
        ttk.Button(g, text="💾  Lưu JSON backup",
                   command=self.on_save_backup, width=36).grid(row=2, column=0, pady=5)

        self.lbl_out = ttk.Label(f, text="", style="Sub.TLabel", wraplength=760)
        self.lbl_out.pack(anchor="w", pady=(14, 0))

    # ---------- Tiện ích luồng / log ----------
    def log(self, msg):
        self.log_q.put(str(msg))

    def _drain_log(self):
        try:
            while True:
                msg = self.log_q.get_nowait()
                self.txt_log.insert("end", msg + "\n")
                self.txt_log.see("end")
        except queue.Empty:
            pass
        self.after(100, self._drain_log)

    def _set_busy(self, busy, status=""):
        self._busy = busy
        self.status.config(text=status or ("Đang chạy..." if busy else "Sẵn sàng"))
        state = "disabled" if busy else "normal"
        self.btn_run.config(state=state)
        self.btn_loadjson.config(state=state)
        self.btn_stop.config(state="normal" if busy else "disabled")
        if busy:
            self.progress.start(12)
        else:
            self.progress.stop()

    def _run_thread(self, target):
        if self._busy:
            return
        self._stop.clear()
        t = threading.Thread(target=target, daemon=True)
        t.start()

    # ---------- Hành động ----------
    def _pick_serpfile(self):
        p = filedialog.askopenfilename(
            title="Chọn file khóa SerpApi",
            filetypes=[("Text", "*.txt"), ("Tất cả", "*.*")])
        if p:
            self.var_serpfile.set(p)
    def _pick_gemfile(self):
        p = filedialog.askopenfilename(
            title="Chọn file khóa Gemini",
            filetypes=[("Text", "*.txt"), ("Tất cả", "*.*")])
        if p:
            self.var_gemfile.set(p)
    def on_stop(self):
        self._stop.set()
        self.log("⏹ Đang yêu cầu dừng...")
        
    def on_run_pipeline(self):
        keyword = self.var_kw.get().strip()
        serpfile = self.var_serpfile.get().strip()
        if not keyword:
            messagebox.showwarning("Thiếu từ khóa", "Nhập từ khóa cần tra cứu.")
            return
        if not serpfile or not os.path.exists(serpfile):
            messagebox.showwarning("Thiếu khóa", "Chọn file khóa SerpApi ở tab Cấu hình.")
            return
        # đọc khóa đầu tiên (bỏ dòng trống); dùng khóa này để gọi Google
        try:
            with open(serpfile, encoding="utf-8") as fp:
                serp_keys = [ln.strip() for ln in fp if ln.strip()]
        except Exception as e:
            messagebox.showerror("Lỗi đọc file khóa", str(e))
            return
        if not serp_keys:
            messagebox.showwarning("File rỗng", "File khóa SerpApi không có khóa nào.")
            return
        serp = serp_keys[0]
        def work():
            self._set_busy(True, "Đang tìm Google...")
            try:
                self.log(f"\n=== Tra cứu: {keyword!r} ===")
                organic = fetch_all_pages(
                    keyword, serp,
                    num_pages=self.var_pages.get(),
                    location=self.var_loc.get().strip() or DEFAULT_LOCATION,
                    progress=self.log,
                    stop_flag=self._stop.is_set)
                if not organic:
                    self.log("Không lấy được kết quả nào.")
                    return
                rows = parse_results(organic)
                self.log(f"Đã trích {len(rows)} kết quả từ Google.")

                if self.var_docrawl.get() and not self._stop.is_set():
                    self.log(f"Đang crawl {len(rows)} trang...")
                    rows = crawl_rows(rows, max_workers=self.var_workers.get())
                self.rows = rows
                self.log(f"✔ Hoàn tất. {len(rows)} dòng sẵn sàng để xuất.")
                self.btn_extract.config(state="normal")
            except Exception as e:                   # noqa: BLE001
                self.log(f"LỖI: {e}")
                messagebox.showerror("Lỗi", str(e))
            finally:
                self._set_busy(False)

        self._run_thread(work)

    def on_extract(self):
        if not self.rows:
            messagebox.showinfo("Chưa có dữ liệu", "Hãy chạy tìm + crawl trước.")
            return
        gemfile = self.var_gemfile.get().strip()
        if not gemfile or not os.path.exists(gemfile):
            messagebox.showwarning("Thiếu khóa Gemini",
                                   "Chọn file khóa Gemini ở tab Cấu hình.")
            return

        def work():
            self._set_busy(True, "Đang trích xuất bằng Gemini...")
            try:
                keys = load_api_keys(gemfile)
                self.log(f"\n=== Trích xuất Gemini ({len(keys)} khóa) ===")
                out = extract_all(
                    keys, self.rows, progress=self.log,
                    stop_flag=self._stop.is_set)
                self.extracted = out
                self.log(f"✔ Trích xuất xong: {len(out)} bản ghi có cấu trúc.")
            except Exception as e:                   # noqa: BLE001
                self.log(f"LỖI: {e}")
                messagebox.showerror("Lỗi", str(e))
            finally:
                self._set_busy(False)

        self._run_thread(work)

    def on_load_backup(self):
        p = filedialog.askopenfilename(
            title="Nạp JSON backup", filetypes=[("JSON", "*.json")])
        if not p:
            return
        try:
            with open(p, encoding="utf-8") as f:
                data = json.load(f)
            # backup crawl: list dict có 'noidung' | backup Gemini: list chuỗi/dict
            if data and isinstance(data[0], dict) and "link" in data[0]:
                self.rows = data
                self.btn_extract.config(state="normal")
                self.log(f"Đã nạp {len(data)} dòng crawl từ {os.path.basename(p)}.")
            else:
                recs, raw = load_records([p])
                self.extracted = recs
                self.log(f"Đã nạp {len(recs)} bản ghi báo giá từ {os.path.basename(p)}.")
        except Exception as e:                       # noqa: BLE001
            messagebox.showerror("Lỗi nạp file", str(e))

    def on_export_excel(self):
        if not self.rows:
            messagebox.showinfo("Chưa có dữ liệu", "Cần dữ liệu crawl để xuất Excel.")
            return
        kw = self.var_kw.get().strip() or "ket_qua"
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx", initialfile=f"{safe_name(kw)}_ket_qua.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not p:
            return
        try:
            build_excel(self.rows, kw, p)
            self.lbl_out.config(text=f"✔ Đã xuất Excel: {p}")
            if messagebox.askyesno("Xong", "Đã xuất Excel. Mở thư mục chứa file?"):
                webbrowser.open("file://" + os.path.dirname(p))
        except Exception as e:                       # noqa: BLE001
            messagebox.showerror("Lỗi ghi Excel", str(e))

    def on_export_html(self):
        recs = self.extracted
        if not recs:
            messagebox.showinfo("Chưa có dữ liệu",
                                "Cần dữ liệu đã trích xuất Gemini (hoặc nạp JSON báo giá).")
            return
        kw = self.var_kw.get().strip() or "Báo giá"
        p = filedialog.asksaveasfilename(
            defaultextension=".html", initialfile=f"{safe_name(kw)}.html",
            filetypes=[("HTML", "*.html")])
        if not p:
            return
        try:
            # chuẩn hóa qua html_builder để chắc chắn đúng khung
            norm = [normalize(r) if isinstance(r, dict) else r
                    for r in recs]
            html = build_html(norm, kw, "app")
            with open(p, "w", encoding="utf-8") as f:
                f.write(html)
            self.lbl_out.config(text=f"✔ Đã xuất HTML: {p}")
            if messagebox.askyesno("Xong", "Đã xuất HTML. Mở bằng trình duyệt?"):
                webbrowser.open("file://" + os.path.abspath(p))
        except Exception as e:                       # noqa: BLE001
            messagebox.showerror("Lỗi ghi HTML", str(e))

    def on_save_backup(self):
        data = self.extracted or self.rows
        if not data:
            messagebox.showinfo("Chưa có dữ liệu", "Không có gì để lưu.")
            return
        kw = self.var_kw.get().strip() or "ket_qua"
        p = filedialog.asksaveasfilename(
            defaultextension=".json", initialfile=f"{safe_name(kw)}_backup.json",
            filetypes=[("JSON", "*.json")])
        if not p:
            return
        try:
            with open(p, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=1)
            self.lbl_out.config(text=f"✔ Đã lưu backup: {p} ({len(data)} mục)")
        except Exception as e:                       # noqa: BLE001
            messagebox.showerror("Lỗi lưu", str(e))


if __name__ == "__main__":
    App().mainloop()
